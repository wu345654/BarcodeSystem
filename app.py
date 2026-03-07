from flask import Flask, render_template, request, jsonify, send_from_directory, session, redirect, url_for
from flask_cors import CORS
import os
import time
from datetime import datetime
from database import (
    init_database, OrderModel, OrderDetailModel, BarcodeModel, ScanRecordModel, LabelTemplateModel,
    UserModel, RoleModel, PermissionModel, UserRoleModel, RolePermissionModel, AuthModel, get_connection,
    DeliveryOrderModel, DeliveryOrderDetailModel
)
from barcode_generator import create_barcodes_for_order, generate_barcode_image
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.fonts import addMapping
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from jinja2 import Template

# 注册中文字体 - 使用本地字体文件
import os
font_path = os.path.join(os.path.dirname(__file__), 'font', 'SIMYOU.TTF')
print(f"字体文件路径: {font_path}")
print(f"字体文件是否存在: {os.path.exists(font_path)}")
try:
    # 注册本地字体文件
    pdfmetrics.registerFont(TTFont('SimYou', font_path))
    addMapping('SimYou', 0, 0, 'SimYou')
    print("本地字体注册成功: SimYou")
except Exception as e:
    print(f"注册本地字体失败: {str(e)}")
    # 尝试使用reportlab的CidFont作为后备
    try:
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
        addMapping('STSong-Light', 0, 0, 'STSong-Light')
        print("后备字体注册成功: STSong-Light")
    except Exception as e:
        print(f"注册后备字体失败: {str(e)}")

# 尝试导入pandas和openpyxl
try:
    import pandas as pd
    import openpyxl
    excel_available = True
except ImportError:
    excel_available = False

# 尝试导入WeasyPrint
try:
    from weasyprint import HTML
    weasyprint_available = True
except (ImportError, OSError):
    weasyprint_available = False

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app)
app.secret_key = 'your-secret-key-here'  # 用于session管理

# 初始化数据库
init_database()


# ==================== 权限控制装饰器 ====================

def login_required(f):
    """登录装饰器"""
    from functools import wraps
    
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    
    return decorated_function


def permission_required(permission_code):
    """权限控制装饰器"""
    from functools import wraps
    
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            
            # 检查用户是否拥有权限
            user_id = session['user_id']
            if not AuthModel.check_user_permission(user_id, permission_code):
                return jsonify({'success': False, 'message': '权限不足'}), 403
            
            return f(*args, **kwargs)
        
        return decorated_function
    
    return decorator


# ==================== 页面路由 ====================

@app.route('/')
@login_required
def index():
    """首页"""
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录页面"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            return render_template('login.html', error='请输入用户名和密码')
        
        # 查找用户
        user = UserModel.get_by_username(username)
        if not user:
            return render_template('login.html', error='用户名或密码错误')
        
        # 验证密码（这里简化处理，实际应该使用密码哈希）
        if user['password'] != password:
            return render_template('login.html', error='用户名或密码错误')
        
        # 登录成功，保存用户信息到session
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['name'] = user['name']
        session['avatar'] = user.get('avatar')
        
        return redirect(url_for('index'))
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    """注销"""
    session.clear()
    return redirect(url_for('login'))


@app.route('/user-profile')
@login_required
def user_profile():
    """个人设置页面"""
    user_id = session['user_id']
    user = UserModel.get_by_id(user_id)
    if not user:
        return redirect(url_for('login'))
    
    # 系统头像列表
    system_avatars = [
        '/static/avatars/avatar1.svg',
        '/static/avatars/avatar2.svg',
        '/static/avatars/avatar3.svg',
        '/static/avatars/avatar4.svg',
        '/static/avatars/avatar5.svg',
        '/static/avatars/avatar6.svg',
        '/static/avatars/avatar7.svg',
        '/static/avatars/avatar8.svg',
        '/static/avatars/avatar9.svg',
        '/static/avatars/avatar10.svg'
    ]
    
    return render_template('user_profile.html', user=user, system_avatars=system_avatars)


@app.route('/orders')
@login_required
def orders_page():
    """订单管理页面"""
    return render_template('orders.html')


@app.route('/scan')
@login_required
def scan_page():
    """条码扫描页面"""
    return render_template('scan.html')


@app.route('/barcodes/<int:order_id>')
@login_required
def barcodes_page(order_id):
    """条码列表页面"""
    return render_template('barcodes.html', order_id=order_id)


@app.route('/print-label/<int:order_id>')
@login_required
def print_label_page(order_id):
    """打印标签页面"""
    return render_template('print_label.html', order_id=order_id)


@app.route('/reports')
@login_required
def reports_page():
    """统计报表页面"""
    return render_template('reports.html')


@app.route('/scan-records')
@login_required
def scan_records_page():
    """扫描记录页面"""
    return render_template('scan_records.html')


@app.route('/label-print')
@login_required
def label_print_page():
    """标签打印页面"""
    # 获取所有订单（不分页）
    orders = OrderModel.get_all(1, 1000)
    return render_template('label_print_2.html', orders=orders)


@app.route('/delivery-order')
@login_required
def delivery_order_page():
    """出库单页面"""
    # 获取所有订单（不分页）
    orders = OrderModel.get_all(1, 1000)
    return render_template('delivery_order.html', orders=orders)


# ==================== 用户管理页面路由 ====================

@app.route('/user-management')
@login_required
def user_management_page():
    """用户管理页面"""
    return render_template('user_management.html')


@app.route('/role-management')
@login_required
def role_management_page():
    """角色管理页面"""
    return render_template('role_management.html')


@app.route('/permission-management')
@login_required
def permission_management_page():
    """权限管理页面"""
    return render_template('permission_management.html')


# ==================== API路由 - 订单管理 ====================

@app.route('/api/orders', methods=['GET'])
@login_required
@permission_required('order.view')
def get_orders():
    """获取所有订单"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    
    orders = OrderModel.get_all(page, page_size)
    # 为每个订单添加扫描统计和明细
    for order in orders:
        stats = BarcodeModel.get_scan_statistics(order['id'])
        order['scan_stats'] = stats
        
        # 获取订单明细
        details = OrderDetailModel.get_by_order(order['id'])
        order['details'] = details
    
    return jsonify({'success': True, 'data': orders})


@app.route('/api/orders/<int:order_id>', methods=['GET'])
@login_required
@permission_required('order.view')
def get_order(order_id):
    """获取单个订单"""
    order = OrderModel.get_by_id(order_id)
    if order:
        stats = BarcodeModel.get_scan_statistics(order_id)
        order['scan_stats'] = stats
        
        # 获取订单明细
        details = OrderDetailModel.get_by_order(order_id)
        order['details'] = details
        
        return jsonify({'success': True, 'data': order})
    return jsonify({'success': False, 'message': '订单不存在'}), 404


@app.route('/api/orders', methods=['POST'])
@login_required
@permission_required('order.create')
def create_order():
    """创建订单"""
    data = request.get_json()
    
    # 验证必填字段
    required_fields = ['work_tag', 'name', 'product']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'success': False, 'message': f'{field} 为必填字段'}), 400
    
    try:
        # 计算订单明细总数量
        details = data.get('details', [])
        total_quantity = sum(detail.get('quantity', 1) for detail in details)
        
        # 创建订单
        order_id = OrderModel.create(
            work_tag=data['work_tag'],
            name=data['name'],
            product=data['product'],
            quantity=total_quantity
        )
        
        # 创建订单明细
        for detail in details:
            OrderDetailModel.create(
                order_id=order_id,
                sequence_no=detail.get('sequence_no', 0),
                product_name=detail.get('product_name', ''),
                color=detail.get('color'),
                thickness=detail.get('thickness'),
                drawing_no=detail.get('drawing_no'),
                quantity=detail.get('quantity', 1)
            )
        
        # 生成条码
        barcodes = create_barcodes_for_order(order_id, details)
        
        return jsonify({
            'success': True,
            'message': '订单创建成功',
            'data': {
                'order_id': order_id,
                'barcodes_count': len(barcodes)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>', methods=['PUT'])
@login_required
@permission_required('order.edit')
def update_order(order_id):
    """更新订单"""
    data = request.get_json()
    
    # 检查订单是否存在
    order = OrderModel.get_by_id(order_id)
    if not order:
        return jsonify({'success': False, 'message': '订单不存在'}), 404
    
    try:
        # 更新订单基本信息
        order_data = {k: v for k, v in data.items() if k != 'details'}
        success = OrderModel.update(order_id, **order_data)
        
        # 更新订单明细
        if 'details' in data:
            # 删除原有明细
            OrderDetailModel.delete_by_order(order_id)
            
            # 计算订单明细总数量
            details = data['details']
            total_quantity = sum(detail.get('quantity', 1) for detail in details)
            
            # 更新订单主表数量
            OrderModel.update(order_id, quantity=total_quantity)
            
            # 创建新明细
            for detail in details:
                OrderDetailModel.create(
                    order_id=order_id,
                    sequence_no=detail.get('sequence_no', 0),
                    product_name=detail.get('product_name', ''),
                    color=detail.get('color'),
                    thickness=detail.get('thickness'),
                    drawing_no=detail.get('drawing_no'),
                    quantity=detail.get('quantity', 1)
                )
            
            # 删除原有条码
            BarcodeModel.delete_by_order(order_id)
            
            # 重新生成条码
            barcodes = create_barcodes_for_order(order_id, details)
        
        if success:
            return jsonify({'success': True, 'message': '订单更新成功'})
        return jsonify({'success': False, 'message': '更新失败'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/<int:order_id>', methods=['DELETE'])
@login_required
@permission_required('order.delete')
def delete_order(order_id):
    """删除订单"""
    order = OrderModel.get_by_id(order_id)
    if not order:
        return jsonify({'success': False, 'message': '订单不存在'}), 404
    
    try:
        success = OrderModel.delete(order_id)
        if success:
            return jsonify({'success': True, 'message': '订单删除成功'})
        return jsonify({'success': False, 'message': '删除失败'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/orders/search', methods=['GET'])
@login_required
@permission_required('order.view')
def search_orders():
    """搜索订单"""
    keyword = request.args.get('keyword', '')
    if not keyword:
        return jsonify({'success': False, 'message': '请输入搜索关键词'}), 400
    
    orders = OrderModel.search(keyword)
    for order in orders:
        stats = BarcodeModel.get_scan_statistics(order['id'])
        order['scan_stats'] = stats
        
        # 获取订单明细
        details = OrderDetailModel.get_by_order(order['id'])
        order['details'] = details
    
    return jsonify({'success': True, 'data': orders})


# ==================== API路由 - 条码管理 ====================

@app.route('/api/orders/<int:order_id>/barcodes', methods=['GET'])
@login_required
@permission_required('order.view')
def get_barcodes(order_id):
    """获取订单的所有条码"""
    order = OrderModel.get_by_id(order_id)
    if not order:
        return jsonify({'success': False, 'message': '订单不存在'}), 404
    
    barcodes = BarcodeModel.get_by_order(order_id)
    # 添加图片路径
    for barcode in barcodes:
        barcode['image_path'] = f"/static/barcodes/order_{order_id}_seq_{barcode['sequence_no']}.png"
    
    return jsonify({
        'success': True,
        'data': {
            'order': order,
            'barcodes': barcodes
        }
    })


@app.route('/api/orders/<int:order_id>/details', methods=['GET'])
@login_required
@permission_required('order.view')
def get_order_details(order_id):
    """获取订单的所有明细"""
    try:
        order = OrderModel.get_by_id(order_id)
        if not order:
            return jsonify({'success': False, 'message': '订单不存在'}), 404
        
        details = OrderDetailModel.get_by_order(order_id)
        return jsonify({'success': True, 'data': details})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/barcodes/<barcode>', methods=['GET'])
@login_required
@permission_required('order.view')
def get_barcode(barcode):
    """根据条码值获取条码信息"""
    barcode_info = BarcodeModel.get_by_barcode(barcode)
    if barcode_info:
        order = OrderModel.get_by_id(barcode_info['order_id'])
        barcode_info['order'] = order
        barcode_info['image_path'] = f"/static/barcodes/{barcode}.png"
        return jsonify({'success': True, 'data': barcode_info})
    return jsonify({'success': False, 'message': '条码不存在'}), 404


# ==================== API路由 - 扫描功能 ====================

@app.route('/api/scan', methods=['POST'])
@login_required
@permission_required('scan.scan')
def scan_barcode():
    """
    扫描条码
    检查条码是否已扫描，防止重复扫描
    """
    data = request.get_json()
    barcode_value = data.get('barcode')
    scanned_by = data.get('scanned_by', '')
    
    if not barcode_value:
        return jsonify({'success': False, 'message': '请输入条码'}), 400
    
    # 查询条码
    barcode_info = BarcodeModel.get_by_barcode(barcode_value)
    
    if not barcode_info:
        # 条码不存在
        ScanRecordModel.create(
            barcode_id=0,
            barcode=barcode_value,
            scan_result='FAILED',
            message='条码不存在',
            scanned_by=scanned_by
        )
        return jsonify({
            'success': False,
            'scan_result': 'NOT_FOUND',
            'message': '条码不存在，请检查条码是否正确'
        }), 404
    
    # 检查是否已扫描
    if barcode_info['is_scanned']:
        # 重复扫描
        ScanRecordModel.create(
            barcode_id=barcode_info['id'],
            barcode=barcode_value,
            scan_result='DUPLICATE',
            message='条码已扫描，不能重复扫描',
            scanned_by=scanned_by
        )
        return jsonify({
            'success': False,
            'scan_result': 'DUPLICATE',
            'message': '⚠️ 该条码已扫描，不能重复扫描！',
            'data': {
                'barcode': barcode_info,
                'order': OrderModel.get_by_id(barcode_info['order_id']),
                'first_scanned_at': barcode_info['scanned_at'],
                'first_scanned_by': barcode_info['scanned_by']
            }
        }), 409
    
    # 正常扫描 - 标记为已扫描
    BarcodeModel.mark_as_scanned(barcode_info['id'], scanned_by)
    
    # 记录扫描成功
    ScanRecordModel.create(
        barcode_id=barcode_info['id'],
        barcode=barcode_value,
        scan_result='SUCCESS',
        message='扫描成功',
        scanned_by=scanned_by
    )
    
    # 获取订单信息
    order = OrderModel.get_by_id(barcode_info['order_id'])
    
    return jsonify({
        'success': True,
        'scan_result': 'SUCCESS',
        'message': '✅ 扫描成功！',
        'data': {
            'barcode': barcode_info,
            'order': order
        }
    })


@app.route('/api/scan-records', methods=['GET'])
@login_required
@permission_required('scan.view')
def get_scan_records():
    """获取扫描记录"""

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    limit = request.args.get('limit', 500, type=int)  # 保持向后兼容
    
    if page and page_size:
        records = ScanRecordModel.get_all(page, page_size)
    else:
        records = ScanRecordModel.get_all(None, None, limit)
    
    return jsonify({'success': True, 'data': records})


# ==================== 静态文件服务 ====================

@app.route('/static/<path:filename>')
def serve_static(filename):
    """提供静态文件"""
    return send_from_directory(app.static_folder, filename)


@app.route('/mic/<path:filename>')
def serve_mic(filename):
    """提供音频文件"""
    mic_dir = os.path.join(os.path.dirname(__file__), 'mic')
    return send_from_directory(mic_dir, filename)


@app.route('/font/<path:filename>')
def serve_font(filename):
    """提供字体文件"""
    font_dir = os.path.join(os.path.dirname(__file__), 'font')
    return send_from_directory(font_dir, filename)


# ==================== 统计信息 ====================

@app.route('/api/statistics', methods=['GET'])
@login_required
@permission_required('report.view')
def get_statistics():
    """获取系统统计信息"""

    conn = get_connection()
    cursor = conn.cursor()
    
    # 订单总数
    cursor.execute('SELECT COUNT(*) as count FROM orders')
    total_orders = cursor.fetchone()['count']
    
    # 条码总数
    cursor.execute('SELECT COUNT(*) as count FROM barcodes')
    total_barcodes = cursor.fetchone()['count']
    
    # 已扫描条码数
    cursor.execute('SELECT COUNT(*) as count FROM barcodes WHERE is_scanned = 1')
    scanned_barcodes = cursor.fetchone()['count']
    
    # 今日扫描数
    cursor.execute('''
        SELECT COUNT(*) as count FROM scan_records 
        WHERE date(scanned_at) = date('now')
    ''')
    today_scans = cursor.fetchone()['count']
    
    conn.close()
    
    return jsonify({
        'success': True,
        'data': {
            'total_orders': total_orders,
            'total_barcodes': total_barcodes,
            'scanned_barcodes': scanned_barcodes,
            'unscanned_barcodes': total_barcodes - scanned_barcodes,
            'today_scans': today_scans
        }
    })


# ==================== API路由 - 标签打印 ====================




@app.route('/api/label-print/generate', methods=['POST'])
@login_required
@permission_required('label.print')
def generate_label_pdf():
    """生成标签PDF（优先使用WeasyPrint，失败时使用reportlab）"""

    # 检查WeasyPrint是否可用
    if not weasyprint_available:
        # WeasyPrint不可用，使用reportlab作为替代
        return generate_label_pdf_fallback()
    
    data = request.get_json()
    order_id = data.get('order_id')
    label_count = data.get('label_count', 10)
    template = data.get('template')
    
    if not order_id:
        return jsonify({'success': False, 'message': '请选择订单'}), 400
    
    # 确保order_id是整数
    try:
        order_id = int(order_id)
    except ValueError:
        return jsonify({'success': False, 'message': '订单ID无效'}), 400
    
    try:
        # 获取订单和条码信息
        print(f"获取订单信息，order_id: {order_id}")
        order = OrderModel.get_by_id(order_id)
        print(f"订单信息: {order}")
        if not order:
            return jsonify({'success': False, 'message': '订单不存在'}), 404
        
        # 获取订单明细
        order_details = OrderDetailModel.get_by_order(order_id)
        print(f"订单明细数量: {len(order_details)}")
        
        barcodes = BarcodeModel.get_by_order(order_id)
        print(f"条码数量: {len(barcodes)}")
        if not barcodes:
            return jsonify({'success': False, 'message': '订单没有条码'}), 404
        
        # 限制标签数量
        barcodes = barcodes[:label_count]
        
        # 创建PDF文件
        pdf_dir = os.path.join(app.static_folder, 'labels')
        os.makedirs(pdf_dir, exist_ok=True)
        
        pdf_filename = f'label2_order_{order_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        pdf_path = os.path.join(pdf_dir, pdf_filename)
        
        # 生成HTML模板
        if template:
            # 尝试解析模板
            try:
                import json
                template_json = json.loads(template)
                
                # 处理模板元素，为每个元素创建Jinja2模板
                template_elements = []
                if 'elements' in template_json:
                    for element in template_json['elements']:
                        element_copy = element.copy()
                        # 为文本元素内容创建Jinja2模板
                        if element_copy.get('type') == 'text' and element_copy.get('content'):
                            # 替换简化的变量语法 {var} 为 Jinja2 语法 {{ var }}
                            content = element_copy['content']
                            import re
                            # 匹配 {variable} 格式的变量
                            content = re.sub(r'\{([^}]+)\}', r'{{ \1 }}', content)
                            element_copy['content_template'] = Template(content)
                        template_elements.append(element_copy)
                
                # 生成自定义模板的HTML
                html_template = '''
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="UTF-8">
                    <title>标签打印</title>
                    <style>
                        @font-face {
                            font-family: 'SimYou';
                            src: url('/font/SIMYOU.TTF') format('truetype');
                        }
                        body, div, p {
                            font-family: 'SimYou', 'STSong', 'SimHei', sans-serif;
                        }
                        .label {
                            width: 80mm;
                            height: 50mm;
                            border: 1px solid black;
                            padding: 0;
                            margin: 5mm;
                            float: left;
                            page-break-inside: avoid;
                            position: relative;
                        }
                    </style>
                </head>
                <body>
                    {% for barcode in barcodes %}
                    <div class="label">
                        {% for element in template_elements %}
                            {% if element.type == 'text' %}
                                <div style="position: absolute; left: {{ element.x }}px; top: {{ element.y }}px; width: {{ element.width }}px; height: {{ element.height }}px; font-size: 12px;">
                                    {% if element.content_template %}
                                        {{ element.content_template.render(order=order, order_cn=order_cn, order_details=order_details, order_details_cn=order_details_cn, first_detail=first_detail, first_detail_cn=first_detail_cn, barcode=barcode)|safe }}
                                    {% else %}
                                        {{ element.content|safe }}
                                    {% endif %}
                                </div>
                            {% elif element.type == 'barcode' %}
                                <div style="position: absolute; left: {{ element.x }}px; top: {{ element.y }}px; width: {{ element.width }}px; height: {{ element.height }}px; text-align: center;">
                                    <img src="{{ barcode.image_path }}" alt="条码" style="max-width: 100%; max-height: 100%;">
                                </div>
                            {% elif element.type == 'line' %}
                                <div style="position: absolute; left: {{ element.x }}px; top: {{ element.y }}px; width: {{ element.width }}px; height: {{ element.height }}px; background-color: #000;"></div>
                            {% elif element.type == 'rect' %}
                                <div style="position: absolute; left: {{ element.x }}px; top: {{ element.y }}px; width: {{ element.width }}px; height: {{ element.height }}px; border: 1px solid #000; background-color: rgba(200, 200, 200, 0.3);"></div>
                            {% endif %}
                        {% endfor %}
                    </div>
                    {% endfor %}
                </body>
                </html>
                '''
                
                # 处理订单数据，添加中文标签
                order_cn = {
                    'id': order.get('id', ''),
                    '工程名称': order.get('work_tag', ''),
                    '备注': order.get('name', ''),
                    '产品': order.get('product', ''),
                    '数量': order.get('quantity', 0),
                    '创建时间': order.get('created_at', ''),
                    '更新时间': order.get('updated_at', '')
                }
                
                # 处理订单明细数据，添加中文标签
                order_details_cn = []
                for detail in order_details:
                    detail_cn = {
                        'id': detail.get('id', ''),
                        '序号': detail.get('sequence_no', ''),
                        '品名': detail.get('product_name', ''),
                        '颜色': detail.get('color', ''),
                        '板厚': detail.get('thickness', ''),
                        '图号': detail.get('drawing_no', ''),
                        '数量': detail.get('quantity', 0)
                    }
                    order_details_cn.append(detail_cn)
                
                # 添加订单明细的第一个元素，方便模板使用
                first_detail = order_details[0] if order_details else {}
                first_detail_cn = order_details_cn[0] if order_details_cn else {}
                
                # 渲染HTML
                jinja_template = Template(html_template)
                render_context = {
                    'order': order,
                    'order_cn': order_cn,
                    'order_details': order_details,
                    'order_details_cn': order_details_cn,
                    'first_detail': first_detail,
                    'first_detail_cn': first_detail_cn,
                    'barcodes': [{
                        'barcode': barcode['barcode'],
                        'sequence_no': barcode['sequence_no'],
                        'image_path': f'/static/barcodes/order_{order_id}_seq_{barcode["sequence_no"]}.png'
                    } for barcode in barcodes],
                    'template_elements': template_elements
                }
                html_content = jinja_template.render(**render_context)
                
            except json.JSONDecodeError:
                # 模板解析失败，使用默认模板
                html_template = '''
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="UTF-8">
                    <title>标签打印</title>
                    <style>
                        @font-face {
                            font-family: 'SimYou';
                            src: url('/font/SIMYOU.TTF') format('truetype');
                        }
                        body, div, p {
                            font-family: 'SimYou', 'STSong', 'SimHei', sans-serif;
                        }
                        .label {
                            width: 80mm;
                            height: 50mm;
                            border: 1px solid black;
                            padding: 5mm;
                            margin: 5mm;
                            float: left;
                            page-break-inside: avoid;
                        }
                        .barcode {
                            text-align: center;
                            margin-bottom: 5mm;
                        }
                        .barcode img {
                            max-width: 100%;
                            height: auto;
                        }
                        .info {
                            font-size: 12px;
                            line-height: 1.2;
                        }
                    </style>
                </head>
                <body>
                    {% for barcode in barcodes %}
                    <div class="label">
                        <div class="barcode">
                            <img src="{{ barcode.image_path }}" alt="条码">
                            <div>{{ barcode.barcode }}</div>
                        </div>
                        <div class="info">
                            <p>工程名称: {{ order.work_tag }}</p>
                            <p>订单名称: {{ order.name }}</p>
                            <p>产品名称: {{ order.product }}</p>
                            <p>颜色: {{ order.color or '-' }}</p>
                            <p>图号: {{ order.drawing_no or '-' }}</p>
                            <p>序号: {{ barcode.sequence_no }}</p>
                        </div>
                    </div>
                    {% endfor %}
                </body>
                </html>
                '''
                
                # 处理订单数据，添加中文标签
                order_cn = {
                    'id': order.get('id', ''),
                    '工程名称': order.get('work_tag', ''),
                    '备注': order.get('name', ''),
                    '产品': order.get('product', ''),
                    '数量': order.get('quantity', 0),
                    '创建时间': order.get('created_at', ''),
                    '更新时间': order.get('updated_at', '')
                }
                
                # 处理订单明细数据，添加中文标签
                order_details_cn = []
                for detail in order_details:
                    detail_cn = {
                        'id': detail.get('id', ''),
                        '序号': detail.get('sequence_no', ''),
                        '品名': detail.get('product_name', ''),
                        '颜色': detail.get('color', ''),
                        '板厚': detail.get('thickness', ''),
                        '图号': detail.get('drawing_no', ''),
                        '数量': detail.get('quantity', 0)
                    }
                    order_details_cn.append(detail_cn)
                
                # 添加订单明细的第一个元素，方便模板使用
                first_detail = order_details[0] if order_details else {}
                first_detail_cn = order_details_cn[0] if order_details_cn else {}
                
                # 渲染HTML
                jinja_template = Template(html_template)
                render_context = {
                    'order': order,
                    'order_cn': order_cn,
                    'order_details': order_details,
                    'order_details_cn': order_details_cn,
                    'barcodes': [{
                        'barcode': barcode['barcode'],
                        'sequence_no': barcode['sequence_no'],
                        'image_path': f'/static/barcodes/order_{order_id}_seq_{barcode["sequence_no"]}.png'
                    } for barcode in barcodes]
                }
                html_content = jinja_template.render(**render_context)
        else:
            # 使用默认模板
            html_template = '''
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>标签打印</title>
                <style>
                    .label {
                        width: 80mm;
                        height: 50mm;
                        border: 1px solid black;
                        padding: 5mm;
                        margin: 5mm;
                        float: left;
                        page-break-inside: avoid;
                    }
                    .barcode {
                        text-align: center;
                        margin-bottom: 5mm;
                    }
                    .barcode img {
                        max-width: 100%;
                        height: auto;
                    }
                    .info {
                        font-size: 12px;
                        line-height: 1.2;
                    }
                </style>
            </head>
            <body>
                {% for barcode in barcodes %}
                <div class="label">
                    <div class="barcode">
                        <img src="{{ barcode.image_path }}" alt="条码">
                        <div>{{ barcode.barcode }}</div>
                    </div>
                    <div class="info">
                        <p>工程名称: {{ order.work_tag }}</p>
                        <p>订单名称: {{ order.name }}</p>
                        <p>产品名称: {{ order.product }}</p>
                        <p>颜色: {{ order.color or '-' }}</p>
                        <p>图号: {{ order.drawing_no or '-' }}</p>
                        <p>序号: {{ barcode.sequence_no }}</p>
                    </div>
                </div>
                {% endfor %}
            </body>
            </html>
            '''
            
            # 处理订单数据，添加中文标签
            order_cn = {
                'id': order.get('id', ''),
                '工程名称': order.get('work_tag', ''),
                '备注': order.get('name', ''),
                '产品': order.get('product', ''),
                '数量': order.get('quantity', 0),
                '创建时间': order.get('created_at', ''),
                '更新时间': order.get('updated_at', '')
            }
            
            # 处理订单明细数据，添加中文标签
            order_details_cn = []
            for detail in order_details:
                detail_cn = {
                    'id': detail.get('id', ''),
                    '序号': detail.get('sequence_no', ''),
                    '品名': detail.get('product_name', ''),
                    '颜色': detail.get('color', ''),
                    '板厚': detail.get('thickness', ''),
                    '图号': detail.get('drawing_no', ''),
                    '数量': detail.get('quantity', 0)
                }
                order_details_cn.append(detail_cn)
            
            # 渲染HTML
            jinja_template = Template(html_template)
            render_context = {
                'order': order,
                'order_cn': order_cn,
                'order_details': order_details,
                'order_details_cn': order_details_cn,
                'barcodes': [{
                    'barcode': barcode['barcode'],
                    'sequence_no': barcode['sequence_no'],
                    'image_path': f'/static/barcodes/order_{order_id}_seq_{barcode["sequence_no"]}.png'
                } for barcode in barcodes]
            }
            html_content = jinja_template.render(**render_context)
        
        # 使用WeasyPrint生成PDF
        HTML(string=html_content).write_pdf(pdf_path)
        
        # 返回PDF路径
        pdf_url = f'/static/labels/{pdf_filename}'
        return jsonify({
            'success': True,
            'data': {
                'pdf_path': pdf_path,
                'pdf_url': pdf_url
            }
        })
    except Exception as e:
        print(f"生成标签PDF失败: {str(e)}")
        return jsonify({'success': False, 'message': f'生成标签PDF失败: {str(e)}'}), 500


# ==================== API路由 - 出库单 ====================

@app.route('/api/delivery-order/generate', methods=['POST'])
@login_required
@permission_required('delivery.generate')
def generate_delivery_order():
    """生成出库单Excel"""

    # 检查pandas和openpyxl是否可用
    if not excel_available:
        return jsonify({'success': False, 'message': 'pandas和openpyxl未安装，请先安装这些依赖'}), 500
    
    data = request.get_json()
    order_id = data.get('order_id')
    
    if not order_id:
        return jsonify({'success': False, 'message': '请选择订单'}), 400
    
    # 确保order_id是整数
    try:
        order_id = int(order_id)
    except ValueError:
        return jsonify({'success': False, 'message': '订单ID无效'}), 400
    
    try:
        # 获取订单和条码信息
        order = OrderModel.get_by_id(order_id)
        if not order:
            return jsonify({'success': False, 'message': '订单不存在'}), 404
        
        barcodes = BarcodeModel.get_by_order(order_id)
        if not barcodes:
            return jsonify({'success': False, 'message': '订单没有条码'}), 404
        
        # 创建Excel文件
        excel_dir = os.path.join(app.static_folder, 'delivery_orders')
        os.makedirs(excel_dir, exist_ok=True)
        
        excel_filename = f'delivery_order_{order_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        excel_path = os.path.join(excel_dir, excel_filename)
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '产品销售发货单'
        
        # 设置列宽，与出库单明细保持一致
        ws.column_dimensions['A'].width = 8  # 序号
        ws.column_dimensions['B'].width = 10  # 楼号
        ws.column_dimensions['C'].width = 12  # 图号
        ws.column_dimensions['D'].width = 15  # 品名
        ws.column_dimensions['E'].width = 8  # 宽
        ws.column_dimensions['F'].width = 8  # 板厚
        ws.column_dimensions['G'].width = 8  # 计量
        ws.column_dimensions['H'].width = 8  # 数量
        ws.column_dimensions['I'].width = 8  # 单件
        ws.column_dimensions['J'].width = 8  # 总量
        ws.column_dimensions['K'].width = 12  # 单件开槽(米)
        ws.column_dimensions['L'].width = 12  # 总开槽(米)
        
        # 标题
        ws.merge_cells('A1:L1')
        ws['A1'] = '产品销售发货单'
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=16)
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        # 发货单信息
        ws['A2'] = '购货单位:'
        ws['B2'] = order.get('name', '')
        ws['F2'] = '发货单号:'
        ws['G2'] = f'0000{order_id}'
        
        ws['A3'] = '项目名称:'
        ws['B3'] = order.get('work_tag', '')
        ws['F3'] = '合同编号:'
        ws['G3'] = ''
        
        ws['A4'] = '送货地址:'
        ws['B4'] = ''
        ws['F4'] = '发货面积:'
        ws['G4'] = ''
        
        ws['A5'] = '收货人:'
        ws['B5'] = ''
        ws['C5'] = '电话:'
        ws['D5'] = ''
        ws['F5'] = '司机电话:'
        ws['G5'] = ''
        
        ws['A6'] = '承运人:'
        ws['B6'] = ''
        ws['C6'] = '车牌:'
        ws['D6'] = ''
        
        # 表头，与出库单明细保持一致
        headers = ['序号', '楼号', '图号', '品名', '宽', '板厚', '计量', '数量', '单件', '总量', '单件开槽(米)', '总开槽(米)']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=i)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        # 数据行
        total_quantity = 0
        total_area = 0
        total_packages = 0
        
        for i, barcode in enumerate(barcodes, 9):
            ws.cell(row=i, column=1, value=i-8).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=2, value='').border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=3, value=order.get('drawing_no', '')).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=4, value=order.get('product', '')).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=5, value='').border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=6, value='').border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=7, value='').border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=8, value=1).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=9, value='').border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=10, value='').border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=11, value='').border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            ws.cell(row=i, column=12, value='').border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            
            total_quantity += 1
        
        # 合计行
        total_row = len(barcodes) + 9
        ws.cell(row=total_row, column=1, value='合计').font = openpyxl.styles.Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_quantity).font = openpyxl.styles.Font(bold=True)
        
        # 签字栏
        sign_row = total_row + 2
        ws.cell(row=sign_row, column=1, value='收货人签字:')
        ws.cell(row=sign_row, column=3, value='承运人签字:')
        ws.cell(row=sign_row, column=5, value='发货人签字:')
        
        ws.cell(row=sign_row+1, column=1, value='日期:')
        ws.cell(row=sign_row+1, column=3, value='日期:')
        ws.cell(row=sign_row+1, column=5, value='日期:')
        
        # 备注栏
        note_row = sign_row + 3
        ws.merge_cells(f'A{note_row}:L{note_row}')
        ws.cell(row=note_row, column=1, value='备注:').font = openpyxl.styles.Font(bold=True)
        
        note_content = [
            '1、本发货单由经办人和收货人签字具有法律效力。',
            '2、货到工地，请轻拿轻放，在搬运过程中出现的损伤、划伤，我公司概不负责；如卸货前发现损坏，划伤',
            '3、收货单位当面清点货物规格、数量、颜色并签字确认。如有异议，自收货之日二日内书面形式通知我公'
        ]
        
        for i, note in enumerate(note_content, note_row+1):
            ws.merge_cells(f'A{i}:L{i}')
            ws.cell(row=i, column=1, value=note)
        
        # 制单信息
        info_row = note_row + len(note_content) + 2
        ws.cell(row=info_row, column=1, value='制单员:')
        ws.cell(row=info_row, column=3, value='审核员:')
        ws.cell(row=info_row, column=5, value='打印员:')
        
        ws.cell(row=info_row+1, column=1, value='制单日期:')
        ws.cell(row=info_row+1, column=3, value='审核日期:')
        ws.cell(row=info_row+1, column=5, value='打印日期:')
        
        # 页码
        page_row = info_row + 2
        ws.merge_cells(f'A{page_row}:L{page_row}')
        ws.cell(row=page_row, column=1, value=f'第 1 页 共 1 页').alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # 保存Excel文件
        wb.save(excel_path)
        
        # 返回Excel路径
        excel_url = f'/static/delivery_orders/{excel_filename}'
        return jsonify({
            'success': True,
            'data': {
                'excel_path': excel_path,
                'excel_url': excel_url
            }
        })
    except Exception as e:
        print(f"生成出库单失败: {str(e)}")
        return jsonify({'success': False, 'message': f'生成出库单失败: {str(e)}'}), 500


# ==================== API路由 - 出库单管理 ====================

@app.route('/api/delivery-orders', methods=['GET'])
@login_required
@permission_required('delivery.view')
def get_delivery_orders():
    """获取出库单列表"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 10))
        search = request.args.get('search', '')

        # 获取出库单列表
        delivery_orders = DeliveryOrderModel.get_all(page, page_size)
        total = DeliveryOrderModel.get_count()

        # 如果有搜索条件，进行过滤
        if search:
            delivery_orders = [
                order for order in delivery_orders
                if search.lower() in (order.get('delivery_no') or '').lower()
                or search.lower() in (order.get('customer_name') or '').lower()
                or search.lower() in (order.get('contract_no') or '').lower()
                or search.lower() in (order.get('project_name') or '').lower()
            ]
            total = len(delivery_orders)

        return jsonify({
            'success': True,
            'data': {
                'items': delivery_orders,
                'total': total,
                'page': page,
                'page_size': page_size
            }
        })
    except Exception as e:
        print(f"获取出库单列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取出库单列表失败: {str(e)}'}), 500


@app.route('/api/delivery-orders/generate-no', methods=['GET'])
@login_required
def generate_delivery_no():
    """生成出库单号"""
    try:
        delivery_no = DeliveryOrderModel.generate_delivery_no()
        return jsonify({
            'success': True,
            'data': {
                'delivery_no': delivery_no
            }
        })
    except Exception as e:
        print(f"生成出库单号失败: {str(e)}")
        return jsonify({'success': False, 'message': f'生成出库单号失败: {str(e)}'}), 500


@app.route('/api/delivery-orders/<int:delivery_order_id>', methods=['GET'])
@login_required
@permission_required('delivery.view')
def get_delivery_order(delivery_order_id):
    """获取出库单详情"""
    try:
        delivery_order = DeliveryOrderModel.get_by_id(delivery_order_id)
        if not delivery_order:
            return jsonify({'success': False, 'message': '出库单不存在'}), 404

        # 获取出库单明细
        details = DeliveryOrderDetailModel.get_by_delivery_order(delivery_order_id)

        return jsonify({
            'success': True,
            'data': {
                'delivery_order': delivery_order,
                'details': details
            }
        })
    except Exception as e:
        print(f"获取出库单详情失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取出库单详情失败: {str(e)}'}), 500


@app.route('/api/delivery-orders', methods=['POST'])
@login_required
@permission_required('delivery.create')
def create_delivery_order():
    """创建出库单"""
    data = request.get_json()

    # 验证必填项
    delivery_no = data.get('delivery_no')
    customer_name = data.get('customer_name')

    if not delivery_no or not customer_name:
        return jsonify({'success': False, 'message': '出库单号和购货单位为必填项'}), 400

    try:
        # 创建出库单表头
        delivery_order_id = DeliveryOrderModel.create(
            delivery_no=delivery_no,
            customer_name=customer_name,
            contract_no=data.get('contract_no'),
            project_name=data.get('project_name'),
            delivery_quantity=data.get('delivery_quantity', 0),
            delivery_address=data.get('delivery_address'),
            delivery_area=data.get('delivery_area'),
            receiver_name=data.get('receiver_name'),
            receiver_phone=data.get('receiver_phone'),
            carrier_name=data.get('carrier_name'),
            plate_number=data.get('plate_number'),
            driver_phone=data.get('driver_phone'),
            status=data.get('status', 'draft')
        )

        # 创建出库单明细
        details = data.get('details', [])
        for detail in details:
            DeliveryOrderDetailModel.create(
                delivery_order_id=delivery_order_id,
                sequence_no=detail.get('sequence_no', 1),
                order_detail_id=detail.get('order_detail_id'),
                building_no=detail.get('building_no'),
                drawing_no=detail.get('drawing_no'),
                product_name=detail.get('product_name'),
                width=detail.get('width'),
                thickness=detail.get('thickness'),
                unit=detail.get('unit'),
                quantity=detail.get('quantity', 0),
                single_weight=detail.get('single_weight'),
                total_weight=detail.get('total_weight'),
                single_groove=detail.get('single_groove'),
                total_groove=detail.get('total_groove')
            )

        return jsonify({
            'success': True,
            'message': '出库单创建成功',
            'data': {
                'delivery_order_id': delivery_order_id
            }
        })
    except Exception as e:
        print(f"创建出库单失败: {str(e)}")
        return jsonify({'success': False, 'message': f'创建出库单失败: {str(e)}'}), 500


@app.route('/api/delivery-orders/<int:delivery_order_id>', methods=['PUT'])
@login_required
@permission_required('delivery.edit')
def update_delivery_order(delivery_order_id):
    """更新出库单"""
    data = request.get_json()

    # 检查出库单是否存在
    delivery_order = DeliveryOrderModel.get_by_id(delivery_order_id)
    if not delivery_order:
        return jsonify({'success': False, 'message': '出库单不存在'}), 404

    try:
        # 更新表头
        DeliveryOrderModel.update(
            delivery_order_id,
            customer_name=data.get('customer_name'),
            contract_no=data.get('contract_no'),
            project_name=data.get('project_name'),
            delivery_quantity=data.get('delivery_quantity'),
            delivery_address=data.get('delivery_address'),
            delivery_area=data.get('delivery_area'),
            receiver_name=data.get('receiver_name'),
            receiver_phone=data.get('receiver_phone'),
            carrier_name=data.get('carrier_name'),
            plate_number=data.get('plate_number'),
            driver_phone=data.get('driver_phone'),
            status=data.get('status')
        )

        # 删除原有明细
        DeliveryOrderDetailModel.delete_by_delivery_order(delivery_order_id)

        # 重新创建明细
        details = data.get('details', [])
        for detail in details:
            DeliveryOrderDetailModel.create(
                delivery_order_id=delivery_order_id,
                sequence_no=detail.get('sequence_no', 1),
                order_detail_id=detail.get('order_detail_id'),
                building_no=detail.get('building_no'),
                drawing_no=detail.get('drawing_no'),
                product_name=detail.get('product_name'),
                width=detail.get('width'),
                thickness=detail.get('thickness'),
                unit=detail.get('unit'),
                quantity=detail.get('quantity', 0),
                single_weight=detail.get('single_weight'),
                total_weight=detail.get('total_weight'),
                single_groove=detail.get('single_groove'),
                total_groove=detail.get('total_groove')
            )

        return jsonify({
            'success': True,
            'message': '出库单更新成功'
        })
    except Exception as e:
        print(f"更新出库单失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新出库单失败: {str(e)}'}), 500


@app.route('/api/delivery-orders/<int:delivery_order_id>', methods=['DELETE'])
@login_required
@permission_required('delivery.delete')
def delete_delivery_order(delivery_order_id):
    """删除出库单"""
    try:
        delivery_order = DeliveryOrderModel.get_by_id(delivery_order_id)
        if not delivery_order:
            return jsonify({'success': False, 'message': '出库单不存在'}), 404

        DeliveryOrderModel.delete(delivery_order_id)

        return jsonify({
            'success': True,
            'message': '出库单删除成功'
        })
    except Exception as e:
        print(f"删除出库单失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除出库单失败: {str(e)}'}), 500


@app.route('/api/delivery-orders/<int:delivery_order_id>/export', methods=['POST'])
@login_required
@permission_required('delivery.export')
def export_delivery_order(delivery_order_id):
    """导出出库单为Excel"""
    try:
        # 检查出库单是否存在
        delivery_order = DeliveryOrderModel.get_by_id(delivery_order_id)
        if not delivery_order:
            return jsonify({'success': False, 'message': '出库单不存在'}), 404

        # 获取出库单明细
        details = DeliveryOrderDetailModel.get_by_delivery_order(delivery_order_id)

        # 创建Excel文件
        excel_dir = os.path.join(app.static_folder, 'delivery_orders')
        os.makedirs(excel_dir, exist_ok=True)

        excel_filename = f'delivery_order_{delivery_order_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        excel_path = os.path.join(excel_dir, excel_filename)

        # 创建工作簿
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "出库单"

        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15

        # 标题
        ws.merge_cells('A1:H1')
        ws['A1'] = '产品销售发货单'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # 表头信息
        ws['A3'] = '出库单号:'
        ws['B3'] = delivery_order['delivery_no']
        ws['D3'] = '购货单位:'
        ws['E3'] = delivery_order['customer_name']

        ws['A4'] = '合同编号:'
        ws['B4'] = delivery_order.get('contract_no', '')
        ws['D4'] = '项目名称:'
        ws['E4'] = delivery_order.get('project_name', '')

        ws['A5'] = '送货地址:'
        ws['B5'] = delivery_order.get('delivery_address', '')
        ws.merge_cells('B5:E5')

        ws['A6'] = '收货人:'
        ws['B6'] = delivery_order.get('receiver_name', '')
        ws['D6'] = '电话:'
        ws['E6'] = delivery_order.get('receiver_phone', '')

        ws['A7'] = '承运人:'
        ws['B7'] = delivery_order.get('carrier_name', '')
        ws['D7'] = '车牌号:'
        ws['E7'] = delivery_order.get('plate_number', '')
        ws['F7'] = '司机电话:'
        ws['G7'] = delivery_order.get('driver_phone', '')

        # 明细表头
        header_row = 9
        headers = ['序号', '楼号', '图号', '品名', '宽', '板厚', '计量', '数量', '单件', '总量', '单件开槽(米)', '总开槽(米)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 明细数据
        for row_idx, detail in enumerate(details, header_row + 1):
            ws.cell(row=row_idx, column=1, value=detail.get('sequence_no', ''))
            ws.cell(row=row_idx, column=2, value=detail.get('building_no', ''))
            ws.cell(row=row_idx, column=3, value=detail.get('drawing_no', ''))
            ws.cell(row=row_idx, column=4, value=detail.get('product_name', ''))
            ws.cell(row=row_idx, column=5, value=detail.get('width', ''))
            ws.cell(row=row_idx, column=6, value=detail.get('thickness', ''))
            ws.cell(row=row_idx, column=7, value=detail.get('unit', ''))
            ws.cell(row=row_idx, column=8, value=detail.get('quantity', ''))
            ws.cell(row=row_idx, column=9, value=detail.get('single_weight', ''))
            ws.cell(row=row_idx, column=10, value=detail.get('total_weight', ''))
            ws.cell(row=row_idx, column=11, value=detail.get('single_groove', ''))
            ws.cell(row=row_idx, column=12, value=detail.get('total_groove', ''))

            # 添加边框
            for col in range(1, 13):
                ws.cell(row=row_idx, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # 保存文件
        wb.save(excel_path)

        excel_url = f'/static/delivery_orders/{excel_filename}'
        return jsonify({
            'success': True,
            'message': '导出成功',
            'data': {
                'excel_url': excel_url
            }
        })
    except Exception as e:
        print(f"导出出库单失败: {str(e)}")
        return jsonify({'success': False, 'message': f'导出出库单失败: {str(e)}'}), 500


# ==================== API路由 - 标签模板 ====================

@app.route('/api/label-templates', methods=['GET'])
@login_required
@permission_required('label.template')
def get_label_templates():
    """获取所有标签模板"""

    try:
        templates = LabelTemplateModel.get_all()
        return jsonify({
            'success': True,
            'data': templates
        })
    except Exception as e:
        print(f"获取模板列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取模板列表失败: {str(e)}'}), 500


@app.route('/api/label-templates', methods=['POST'])
@login_required
@permission_required('label.template')
def create_label_template():
    """创建标签模板"""

    data = request.get_json()
    name = data.get('name')
    template = data.get('template')
    
    if not name or not template:
        return jsonify({'success': False, 'message': '模板名称和内容为必填项'}), 400
    
    try:
        template_id = LabelTemplateModel.create(name, template)
        return jsonify({
            'success': True,
            'message': '模板创建成功',
            'data': {
                'template_id': template_id
            }
        })
    except Exception as e:
        print(f"创建模板失败: {str(e)}")
        return jsonify({'success': False, 'message': f'创建模板失败: {str(e)}'}), 500


@app.route('/api/label-templates/<int:template_id>', methods=['GET'])
@login_required
@permission_required('label.template')
def get_label_template(template_id):
    """获取标签模板详情"""

    try:
        template = LabelTemplateModel.get_by_id(template_id)
        if not template:
            return jsonify({'success': False, 'message': '模板不存在'}), 404
        return jsonify({
            'success': True,
            'data': template
        })
    except Exception as e:
        print(f"获取模板详情失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取模板详情失败: {str(e)}'}), 500


@app.route('/api/user/profile', methods=['POST'])
@login_required
def update_user_profile():
    """更新用户个人设置"""
    try:
        user_id = session['user_id']
        user = UserModel.get_by_id(user_id)
        if not user:
            return jsonify({'success': False, 'message': '用户不存在'}), 404
        
        # 处理表单数据
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        avatar_url = request.form.get('avatar_url')
        
        # 验证数据
        if not name:
            return jsonify({'success': False, 'message': '姓名不能为空'}), 400
        
        if password:
            if password != confirm_password:
                return jsonify({'success': False, 'message': '两次输入的密码不一致'}), 400
        
        # 处理头像上传
        avatar = None
        if 'avatar' in request.files:
            file = request.files['avatar']
            if file and file.filename:
                # 确保avatars目录存在
                avatars_dir = os.path.join(app.static_folder, 'avatars')
                os.makedirs(avatars_dir, exist_ok=True)
                
                # 生成唯一文件名
                filename = f"{user_id}_{int(time.time())}_{file.filename}"
                filepath = os.path.join(avatars_dir, filename)
                file.save(filepath)
                
                # 生成相对路径
                avatar = f"/static/avatars/{filename}"
        elif avatar_url:
            # 使用系统头像
            avatar = avatar_url
        
        # 更新用户信息
        update_data = {'name': name, 'email': email}
        if password:
            update_data['password'] = password
        if avatar:
            update_data['avatar'] = avatar
        
        UserModel.update(user_id, **update_data)
        
        # 更新session中的用户信息
        session['name'] = name
        if avatar:
            session['avatar'] = avatar
        
        return jsonify({
            'success': True,
            'message': '个人设置更新成功',
            'data': {
                'name': name,
                'email': email,
                'avatar': avatar
            }
        })
        
    except Exception as e:
        print(f"更新个人设置失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新个人设置失败: {str(e)}'}), 500


@app.route('/api/user/permissions', methods=['GET'])
@login_required
def get_user_permissions():
    """获取用户权限"""
    try:
        user_id = session['user_id']
        permissions = AuthModel.get_user_permissions(user_id)
        return jsonify({
            'success': True,
            'data': permissions
        })
    except Exception as e:
        print(f"获取用户权限失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取用户权限失败: {str(e)}'}), 500


@app.route('/api/label-templates/<int:template_id>', methods=['PUT'])
@login_required
@permission_required('label.template')
def update_label_template(template_id):
    """更新标签模板"""

    data = request.get_json()
    name = data.get('name')
    template = data.get('template')
    
    if not name or not template:
        return jsonify({'success': False, 'message': '模板名称和内容为必填项'}), 400
    
    try:
        success = LabelTemplateModel.update(template_id, name, template)
        if success:
            return jsonify({'success': True, 'message': '模板更新成功'})
        return jsonify({'success': False, 'message': '模板不存在'}), 404
    except Exception as e:
        print(f"更新模板失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新模板失败: {str(e)}'}), 500


@app.route('/api/label-templates/<int:template_id>', methods=['DELETE'])
@login_required
@permission_required('label.template')
def delete_label_template(template_id):
    """删除标签模板"""

    try:
        success = LabelTemplateModel.delete(template_id)
        if success:
            return jsonify({'success': True, 'message': '模板删除成功'})
        return jsonify({'success': False, 'message': '模板不存在'}), 404
    except Exception as e:
        print(f"删除模板失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除模板失败: {str(e)}'}), 500


# ==================== API路由 - 用户管理 ====================

@app.route('/api/users', methods=['GET'])
@login_required
@permission_required('user.manage')
def get_users():
    """获取所有用户"""

    try:
        users = UserModel.get_all()
        return jsonify({'success': True, 'data': users})
    except Exception as e:
        print(f"获取用户列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取用户列表失败: {str(e)}'}), 500


@app.route('/api/users', methods=['POST'])
@login_required
@permission_required('user.manage')
def create_user():
    """创建用户"""

    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    name = data.get('name')
    email = data.get('email')
    role_ids = data.get('role_ids', [])
    
    if not username or not password or not name:
        return jsonify({'success': False, 'message': '用户名、密码和姓名为必填项'}), 400
    
    try:
        # 创建用户
        user_id = UserModel.create(username, password, name, email)
        
        # 分配角色
        for role_id in role_ids:
            UserRoleModel.create(user_id, role_id)
        
        return jsonify({'success': True, 'message': '用户创建成功', 'data': {'user_id': user_id}})
    except Exception as e:
        print(f"创建用户失败: {str(e)}")
        return jsonify({'success': False, 'message': f'创建用户失败: {str(e)}'}), 500


@app.route('/api/users/<int:user_id>', methods=['GET'])
@login_required
@permission_required('user.manage')
def get_user(user_id):
    """获取用户详情"""

    try:
        user = UserModel.get_by_id(user_id)
        if not user:
            return jsonify({'success': False, 'message': '用户不存在'}), 404
        
        # 获取用户角色
        roles = UserRoleModel.get_roles_by_user(user_id)
        user['roles'] = roles
        
        return jsonify({'success': True, 'data': user})
    except Exception as e:
        print(f"获取用户详情失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取用户详情失败: {str(e)}'}), 500


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
@permission_required('user.manage')
def update_user(user_id):
    """更新用户"""

    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    password = data.get('password')
    role_ids = data.get('role_ids', [])
    
    try:
        # 更新用户信息
        updates = {}
        if name:
            updates['name'] = name
        if email:
            updates['email'] = email
        if password:
            updates['password'] = password
        
        if updates:
            UserModel.update(user_id, **updates)
        
        # 更新角色
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM user_roles WHERE user_id = ?', (user_id,))
        conn.commit()
        conn.close()
        
        for role_id in role_ids:
            UserRoleModel.create(user_id, role_id)
        
        return jsonify({'success': True, 'message': '用户更新成功'})
    except Exception as e:
        print(f"更新用户失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新用户失败: {str(e)}'}), 500


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@permission_required('user.manage')
def delete_user(user_id):
    """删除用户"""

    try:
        success = UserModel.delete(user_id)
        if success:
            return jsonify({'success': True, 'message': '用户删除成功'})
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    except Exception as e:
        print(f"删除用户失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除用户失败: {str(e)}'}), 500


# ==================== API路由 - 角色管理 ====================

@app.route('/api/roles', methods=['GET'])
@login_required
@permission_required('user.manage')
def get_roles():
    """获取所有角色"""

    try:
        roles = RoleModel.get_all()
        return jsonify({'success': True, 'data': roles})
    except Exception as e:
        print(f"获取角色列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取角色列表失败: {str(e)}'}), 500


@app.route('/api/roles', methods=['POST'])
@login_required
@permission_required('user.manage')
def create_role():
    """创建角色"""

    data = request.get_json()
    name = data.get('name')
    description = data.get('description')
    permission_ids = data.get('permission_ids', [])
    
    if not name:
        return jsonify({'success': False, 'message': '角色名称为必填项'}), 400
    
    try:
        # 创建角色
        role_id = RoleModel.create(name, description)
        
        # 分配权限
        for permission_id in permission_ids:
            RolePermissionModel.create(role_id, permission_id)
        
        return jsonify({'success': True, 'message': '角色创建成功', 'data': {'role_id': role_id}})
    except Exception as e:
        print(f"创建角色失败: {str(e)}")
        return jsonify({'success': False, 'message': f'创建角色失败: {str(e)}'}), 500


@app.route('/api/roles/<int:role_id>', methods=['GET'])
@login_required
@permission_required('user.manage')
def get_role(role_id):
    """获取角色详情"""

    try:
        role = RoleModel.get_by_id(role_id)
        if not role:
            return jsonify({'success': False, 'message': '角色不存在'}), 404
        
        # 获取角色权限
        permissions = RolePermissionModel.get_permissions_by_role(role_id)
        role['permissions'] = permissions
        
        return jsonify({'success': True, 'data': role})
    except Exception as e:
        print(f"获取角色详情失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取角色详情失败: {str(e)}'}), 500


@app.route('/api/roles/<int:role_id>', methods=['PUT'])
@login_required
@permission_required('user.manage')
def update_role(role_id):
    """更新角色"""

    data = request.get_json()
    name = data.get('name')
    description = data.get('description')
    permission_ids = data.get('permission_ids', [])
    
    try:
        # 更新角色信息
        conn = get_connection()
        cursor = conn.cursor()
        if name:
            cursor.execute('UPDATE roles SET name = ?, description = ? WHERE id = ?', (name, description, role_id))
            conn.commit()
        conn.close()
        
        # 更新权限
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM role_permissions WHERE role_id = ?', (role_id,))
        conn.commit()
        conn.close()
        
        for permission_id in permission_ids:
            RolePermissionModel.create(role_id, permission_id)
        
        return jsonify({'success': True, 'message': '角色更新成功'})
    except Exception as e:
        print(f"更新角色失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新角色失败: {str(e)}'}), 500


@app.route('/api/roles/<int:role_id>', methods=['DELETE'])
@login_required
@permission_required('user.manage')
def delete_role(role_id):
    """删除角色"""

    try:
        # 先删除角色权限关联
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM role_permissions WHERE role_id = ?', (role_id,))
        cursor.execute('DELETE FROM user_roles WHERE role_id = ?', (role_id,))
        conn.commit()
        conn.close()
        
        # 删除角色
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM roles WHERE id = ?', (role_id,))
        affected = cursor.rowcount
        conn.commit()
        conn.close()
        
        if affected > 0:
            return jsonify({'success': True, 'message': '角色删除成功'})
        return jsonify({'success': False, 'message': '角色不存在'}), 404
    except Exception as e:
        print(f"删除角色失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除角色失败: {str(e)}'}), 500


# ==================== API路由 - 权限管理 ====================

@app.route('/api/permissions', methods=['GET'])
@login_required
@permission_required('user.manage')
def get_permissions():
    """获取所有权限"""

    try:
        permissions = PermissionModel.get_all()
        return jsonify({'success': True, 'data': permissions})
    except Exception as e:
        print(f"获取权限列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取权限列表失败: {str(e)}'}), 500


@app.route('/api/permissions', methods=['POST'])
@login_required
@permission_required('user.manage')
def create_permission():
    """创建权限"""

    data = request.get_json()
    name = data.get('name')
    code = data.get('code')
    description = data.get('description')
    
    if not name or not code:
        return jsonify({'success': False, 'message': '权限名称和代码为必填项'}), 400
    
    try:
        permission_id = PermissionModel.create(name, code, description)
        return jsonify({'success': True, 'message': '权限创建成功', 'data': {'permission_id': permission_id}})
    except Exception as e:
        print(f"创建权限失败: {str(e)}")
        return jsonify({'success': False, 'message': f'创建权限失败: {str(e)}'}), 500


@app.route('/api/permissions/<int:permission_id>', methods=['GET'])
@login_required
@permission_required('user.manage')
def get_permission(permission_id):
    """获取权限详情"""

    try:
        permission = PermissionModel.get_by_id(permission_id)
        if not permission:
            return jsonify({'success': False, 'message': '权限不存在'}), 404
        return jsonify({'success': True, 'data': permission})
    except Exception as e:
        print(f"获取权限详情失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取权限详情失败: {str(e)}'}), 500


@app.route('/api/permissions/<int:permission_id>', methods=['PUT'])
@login_required
@permission_required('user.manage')
def update_permission(permission_id):
    """更新权限"""

    data = request.get_json()
    name = data.get('name')
    code = data.get('code')
    description = data.get('description')
    
    try:
        conn = get_connection()
        cursor = conn.cursor()
        if name and code:
            cursor.execute('UPDATE permissions SET name = ?, code = ?, description = ? WHERE id = ?', (name, code, description, permission_id))
            conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '权限更新成功'})
    except Exception as e:
        print(f"更新权限失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新权限失败: {str(e)}'}), 500


@app.route('/api/permissions/<int:permission_id>', methods=['DELETE'])
@login_required
@permission_required('user.manage')
def delete_permission(permission_id):
    """删除权限"""

    try:
        # 先删除角色权限关联
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM role_permissions WHERE permission_id = ?', (permission_id,))
        conn.commit()
        conn.close()
        
        # 删除权限
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM permissions WHERE id = ?', (permission_id,))
        affected = cursor.rowcount
        conn.commit()
        conn.close()
        
        if affected > 0:
            return jsonify({'success': True, 'message': '权限删除成功'})
        return jsonify({'success': False, 'message': '权限不存在'}), 404
    except Exception as e:
        print(f"删除权限失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除权限失败: {str(e)}'}), 500





def generate_label_pdf_fallback():
    """生成标签PDF（使用reportlab作为替代）"""
    data = request.get_json()
    order_id = data.get('order_id')
    label_count = data.get('label_count', 10)
    template = data.get('template')
    
    if not order_id:
        return jsonify({'success': False, 'message': '请选择订单'}), 400
    
    # 确保order_id是整数
    try:
        order_id = int(order_id)
    except ValueError:
        return jsonify({'success': False, 'message': '订单ID无效'}), 400
    
    try:
        # 获取订单和条码信息
        print(f"获取订单信息，order_id: {order_id}")
        order = OrderModel.get_by_id(order_id)
        print(f"订单信息: {order}")
        if not order:
            return jsonify({'success': False, 'message': '订单不存在'}), 404
        
        # 获取订单明细
        order_details = OrderDetailModel.get_by_order(order_id)
        print(f"订单明细数量: {len(order_details)}")
        
        barcodes = BarcodeModel.get_by_order(order_id)
        print(f"条码数量: {len(barcodes)}")
        if not barcodes:
            return jsonify({'success': False, 'message': '订单没有条码'}), 404
        
        # 限制标签数量
        barcodes = barcodes[:label_count]
        
        # 创建PDF文件
        pdf_dir = os.path.join(app.static_folder, 'labels')
        os.makedirs(pdf_dir, exist_ok=True)
        
        pdf_filename = f'label2_order_{order_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        pdf_path = os.path.join(pdf_dir, pdf_filename)
        
        # 设置页面大小为80mm x 50mm
        page_width = 80 * mm
        page_height = 50 * mm
        
        # 创建PDF画布
        c = canvas.Canvas(pdf_path, pagesize=(page_width, page_height))
        
        # 标签尺寸 (80mm x 50mm)，对应预览的300px x 187px
        label_width = 80 * mm
        label_height = 50 * mm
        
        # 每个标签显示为一页
        labels_per_page = 1
        
        # 生成标签
        for i, barcode in enumerate(barcodes):
            # 每个标签单独一页
            if i > 0:
                c.showPage()
            
            # 设置边距为0 - 使用reportlab正确的方法
            c.setPageSize((page_width, page_height))
            # reportlab的Canvas没有setLeftMargin等方法，边距通过坐标直接控制
            
            # 坐标原点设置为页面左上角
            x = 0
            y = page_height
            
            # 绘制标签边框
            c.rect(x, y, label_width, label_height)
            
            # 绘制内容
            # 为默认模板设置字体，确保中文显示正常
            try:
                c.setFont('SimYou', 10)
            except:
                try:
                    c.setFont('STSong-Light', 10)
                except:
                    try:
                        c.setFont('SimHei', 10)
                    except:
                        try:
                            c.setFont('Heiti TC', 10)
                        except:
                            c.setFont('Helvetica', 10)
            
            if template:
                # 尝试解析模板
                try:
                    import json
                    template_json = json.loads(template)
                    
                    # 根据模板中的元素渲染PDF
                    if 'elements' in template_json:
                        # 检查是否有条码元素
                        has_barcode_element = any(element.get('type') == 'barcode' for element in template_json['elements'])
                        
                        # 如果没有条码元素，添加默认条码图片
                        if not has_barcode_element:
                            barcode_image_path = os.path.join(app.static_folder, 'barcodes', f'order_{order_id}_seq_{barcode["sequence_no"]}.png')
                            if os.path.exists(barcode_image_path):
                                img = ImageReader(barcode_image_path)
                                c.drawImage(img, x + 20 * 0.3527, y + label_height - 80 * 0.3527 - 10, width=150 * 0.3527, height=40 * 0.3527)
                        
                        for element in template_json['elements']:
                            element_type = element.get('type')
                            element_x = element.get('x', 0)
                            element_y = element.get('y', 0)
                            element_content = element.get('content', '')
                            
                            # 渲染内容，替换变量
                            try:
                                # 先进行简单的变量替换，确保即使Jinja2模板失败也能正常工作
                                rendered_content = element_content
                                
                                # 替换order变量
                                rendered_content = rendered_content.replace('{order.id}', str(order.get('id', '')))
                                rendered_content = rendered_content.replace('{order.work_tag}', order.get('work_tag', ''))
                                rendered_content = rendered_content.replace('{order.name}', order.get('name', ''))
                                rendered_content = rendered_content.replace('{order.product}', order.get('product', ''))
                                rendered_content = rendered_content.replace('{order.quantity}', str(order.get('quantity', 0)))
                                rendered_content = rendered_content.replace('{order.created_at}', order.get('created_at', ''))
                                rendered_content = rendered_content.replace('{order.updated_at}', order.get('updated_at', ''))
                                
                                # 替换中文标签变量
                                rendered_content = rendered_content.replace('{order_cn.id}', str(order.get('id', '')))
                                rendered_content = rendered_content.replace('{order_cn.工程名称}', order.get('work_tag', ''))
                                rendered_content = rendered_content.replace('{order_cn.备注}', order.get('name', ''))
                                rendered_content = rendered_content.replace('{order_cn.产品}', order.get('product', ''))
                                rendered_content = rendered_content.replace('{order_cn.数量}', str(order.get('quantity', 0)))
                                rendered_content = rendered_content.replace('{order_cn.创建时间}', order.get('created_at', ''))
                                rendered_content = rendered_content.replace('{order_cn.更新时间}', order.get('updated_at', ''))
                                
                                # 替换订单明细变量
                                if order_details and len(order_details) > 0:
                                    first_detail = order_details[0]
                                    rendered_content = rendered_content.replace('{order_details[0].id}', str(first_detail.get('id', '')))
                                    rendered_content = rendered_content.replace('{order_details[0].sequence_no}', str(first_detail.get('sequence_no', '')))
                                    rendered_content = rendered_content.replace('{order_details[0].product_name}', first_detail.get('product_name', ''))
                                    rendered_content = rendered_content.replace('{order_details[0].color}', first_detail.get('color', ''))
                                    rendered_content = rendered_content.replace('{order_details[0].thickness}', first_detail.get('thickness', ''))
                                    rendered_content = rendered_content.replace('{order_details[0].drawing_no}', first_detail.get('drawing_no', ''))
                                    rendered_content = rendered_content.replace('{order_details[0].quantity}', str(first_detail.get('quantity', 0)))
                                    
                                    # 替换订单明细中文标签变量
                                    rendered_content = rendered_content.replace('{first_detail_cn.id}', str(first_detail.get('id', '')))
                                    rendered_content = rendered_content.replace('{first_detail_cn.序号}', str(first_detail.get('sequence_no', '')))
                                    rendered_content = rendered_content.replace('{first_detail_cn.品名}', first_detail.get('product_name', ''))
                                    rendered_content = rendered_content.replace('{first_detail_cn.颜色}', first_detail.get('color', ''))
                                    rendered_content = rendered_content.replace('{first_detail_cn.板厚}', first_detail.get('thickness', ''))
                                    rendered_content = rendered_content.replace('{first_detail_cn.图号}', first_detail.get('drawing_no', ''))
                                    rendered_content = rendered_content.replace('{first_detail_cn.数量}', str(first_detail.get('quantity', 0)))
                                
                                # 替换barcode变量
                                rendered_content = rendered_content.replace('{barcode}', barcode.get('barcode', ''))
                                rendered_content = rendered_content.replace('{sequence_no}', str(barcode.get('sequence_no', '')))
                            except Exception as e:
                                rendered_content = element_content
                            
                            # 计算元素在PDF中的位置，使用与预览相同的坐标系统
                            # 预览使用300px宽，PDF使用80mm宽，比例约为3.75px/mm
                            scale_factor = 3.75
                            element_pdf_x = x + (element_x / scale_factor) * mm
                            element_pdf_y = y - (element_y / scale_factor) * mm
                            
                            # 根据元素类型渲染
                            if element_type == 'text':
                                # 绘制文本
                                # 设置字体大小，根据元素高度自动调整
                                font_size = min(12, (element.get('height', 30) / 3.75) * 0.8)
                                try:
                                    c.setFont('SimYou', font_size)
                                except:
                                    try:
                                        c.setFont('STSong-Light', font_size)
                                    except:
                                        try:
                                            c.setFont('SimHei', font_size)
                                        except:
                                            try:
                                                c.setFont('Heiti TC', font_size)
                                            except:
                                                c.setFont('Helvetica', font_size)
                                # 绘制文本
                                c.drawString(element_pdf_x, element_pdf_y, rendered_content)
                            elif element_type == 'barcode':
                                # 绘制条码
                                barcode_image_path = os.path.join(app.static_folder, 'barcodes', f'order_{order_id}_seq_{barcode["sequence_no"]}.png')
                                if os.path.exists(barcode_image_path):
                                    img = ImageReader(barcode_image_path)
                                    c.drawImage(img, element_pdf_x, element_pdf_y - (element.get('height', 40) / 3.75) * mm, width=(element.get('width', 120) / 3.75) * mm, height=(element.get('height', 40) / 3.75) * mm)
                            elif element_type == 'line':
                                # 绘制线条
                                c.line(element_pdf_x, element_pdf_y, element_pdf_x + (element.get('width', 100) / 3.75) * mm, element_pdf_y)
                            elif element_type == 'rect':
                                # 绘制矩形
                                c.rect(element_pdf_x, element_pdf_y - (element.get('height', 50) / 3.75) * mm, (element.get('width', 80) / 3.75) * mm, (element.get('height', 50) / 3.75) * mm)
                    else:
                        # 模板格式不正确，使用默认模板
                        c.drawString(x + 5 * mm, y + label_height - 10 * mm, f'模板格式不正确')
                except json.JSONDecodeError:
                    # 不是JSON格式，使用Jinja2模板
                    try:
                        jinja_template = Template(template)
                        # 确保order和barcode变量存在
                        render_context = {
                            'order': order or {},
                            'barcode': barcode or {}
                        }
                        rendered_template = jinja_template.render(**render_context)
                        lines = rendered_template.strip().split('\n')
                        y_offset = y + label_height - 10 * mm
                        for line in lines:
                            if line.strip():
                                c.drawString(x + 5 * mm, y_offset, line.strip())
                                y_offset -= 12
                    except Exception as e:
                        # 模板渲染失败，使用默认模板
                        c.drawString(x + 5 * mm, y - 10 * mm, f'模板渲染失败: {str(e)}')
            else:
                # 使用默认模板
                # 工程名称
                c.drawString(x + 5 * mm, y - 10 * mm, f'工程名称: {order["work_tag"]}')
                # 订单名称
                c.drawString(x + 5 * mm, y - 22 * mm, f'订单名称: {order["name"]}')
                # 产品名称
                c.drawString(x + 5 * mm, y - 34 * mm, f'产品名称: {order["product"]}')
                # 颜色
                c.drawString(x + 5 * mm, y - 46 * mm, f'颜色: {order.get("color", "-")}')
                # 图号
                c.drawString(x + 5 * mm, y - 58 * mm, f'图号: {order.get("drawing_no", "-")}')
                # 条码
                c.drawString(x + 5 * mm, y - 70 * mm, f'条码: {barcode["barcode"]}')
                # 序号
                c.drawString(x + 5 * mm, y - 82 * mm, f'序号: {barcode["sequence_no"]}')
        
        # 保存PDF
        c.save()
        
        # 生成PDF URL
        pdf_url = f'/static/labels/{pdf_filename}'
        
        return jsonify({
            'success': True,
            'data': {
                'pdf_url': pdf_url,
                'pdf_path': pdf_path
            }
        })
    except Exception as e:
        print(f'生成标签失败: {str(e)}')
        return jsonify({'success': False, 'message': f'生成标签失败: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8888)
